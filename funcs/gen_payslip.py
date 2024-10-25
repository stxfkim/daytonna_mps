import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import streamlit as st
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment



def report_gaji_harian(df):
    # Ensure the 'tanggal' column is in datetime format
    df['tanggal'] = pd.to_datetime(df['tanggal'], format='%Y-%m-%d')

    # Group by 'Project', 'nama', and 'jabatan' and pivot the table
    pivot_df = df.pivot_table(index=['nama', 'jabatan','uang_makan','basic_salary','tax_deduction','bpjstk_deduction'], 
                              columns=df['tanggal'].dt.day, 
                              values='billable_hours', 
                              aggfunc='sum').reset_index()
    
    # Create a list of all days (1 to 31)
    all_days = [str(i) for i in range(1, 32)]
    
    # Reindex the pivot table to ensure all days are present
    pivot_df = pivot_df.reindex(columns=['nama', 'jabatan','uang_makan','basic_salary','tax_deduction','bpjstk_deduction'] + list(range(1, 32)), fill_value=0)
    
    # Rename columns
    pivot_df.columns = ['nama', 'jabatan','uang_makan','basic_salary','potongan_pajak','potongan_bpjstk'] + all_days

    #nama, jabatan, days, total jam kerja,total hari kerja, uang_makan,total uang makan, basic salary,
    #jumlah gaji(total basic),
    # pendapatan kotor (subtotal), potongan spatu/helm(isi 0)
    #potongan pph (tax_deduction),
    #subtotal ()
    #bpjs deduction

    pivot_df['total_jam_kerja'] = pivot_df.loc[:, '1':'31'].sum(axis=1)
    pivot_df['total_hari_kerja'] = (pivot_df.loc[:, '1':'31'] > 5).sum(axis=1)
    pivot_df['total_uang_makan'] = pivot_df['uang_makan'] * pivot_df['total_hari_kerja']
    pivot_df['jumlah_gaji'] = pivot_df['total_jam_kerja'] * pivot_df['basic_salary']
    pivot_df['pendapatan_kotor'] = pivot_df['jumlah_gaji'] + pivot_df['total_uang_makan']
    pivot_df['potongan_sepatu_helm'] = 0
    pivot_df['kasbon'] = 0
    pivot_df['tambahan_lain'] = 0
    
    pivot_df['tmp_plus'] = (pivot_df['jumlah_gaji']+pivot_df['total_uang_makan']+pivot_df['tambahan_lain'])
    pivot_df['tmp_minus'] = (pivot_df['potongan_bpjstk']+pivot_df['potongan_sepatu_helm']+pivot_df['potongan_pajak']+pivot_df['kasbon'])
    pivot_df['subtotal'] = pivot_df['tmp_plus']-pivot_df['tmp_minus']
    
    pivot_df['bpjs_dna'] = pivot_df['jumlah_gaji'].apply(
        lambda x: 0 if x <= 1000000 else
                  165400 if x <= 3000000 else
                  163370 if x <= 5000000 else
                  237176.35
    )
    
    final_df = pivot_df[['nama', 'jabatan'] + all_days + ['total_jam_kerja','total_hari_kerja','uang_makan',
                                                          'total_uang_makan','basic_salary','jumlah_gaji',
                                                          'pendapatan_kotor','potongan_sepatu_helm',
                                                          'potongan_pajak','kasbon','tambahan_lain','potongan_bpjstk',
                                                          'subtotal','bpjs_dna']]
    
    
    return final_df

def generate_payslip(working_hours_df,summary_salary_df,detail_salary_df,periode):
    
    working_hours_df['sheet_name'] = working_hours_df['nik'] + "_" +working_hours_df["nama"].replace(
            " ", "_", regex=True)
    
    working_hours_df['sheet_name'] =working_hours_df['sheet_name'].str.upper()
    
    summary_salary_df['sheet_name'] = summary_salary_df['sheet_name'].str.upper()
    
    working_hours_df['jam_mulai'] = working_hours_df['jam_mulai'].dt.strftime('%H:%M')
    working_hours_df['jam_akhir'] = working_hours_df['jam_akhir'].dt.strftime('%H:%M')

    wb = openpyxl.load_workbook(Path('template/template_slipgaji.xlsx'))
    sheetname_list = summary_salary_df['sheet_name'].unique()
    
    filename = periode.replace(" ", "_")
    
    file_output = Path('output/report_'+filename+".xlsx")
    
    for sheetname in sheetname_list:
        #st.write(sheetname)
        tmp_detail = working_hours_df[working_hours_df['sheet_name']==str(sheetname)]
        tmp_summary = summary_salary_df[summary_salary_df['sheet_name']==str(sheetname)]
        
        sheet = wb.copy_worksheet(wb.active)
        sheet.title = sheetname

        start_row = 5
        start_col = 1  # Column A
        # Write detail working hours
        tmp_detail_table = tmp_detail[['tanggal','jam_mulai', 'jam_akhir','billable_hours']]
        for r_idx, row in enumerate(dataframe_to_rows(tmp_detail_table, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # B42 : Name 
        sheet['B42'].value = tmp_summary['nama'].iloc[0]         
        
        # H4 : Employee Number
        sheet['H4'].value = tmp_summary['nik'].iloc[0]
        # H5 : Name 
        sheet['H5'].value = tmp_summary['nama'].iloc[0] 
        # H6 : Jabatan 
        sheet['H6'].value = tmp_summary['jabatan'].iloc[0] 
        # H7 : Status Keluarga 
        sheet['H7'].value = tmp_summary['status_pajak'].iloc[0] 
        # H8 : NPWP 
        sheet['H8'].value = tmp_summary['npwp'].iloc[0] 
        # H9 : Periode Gaji 
        sheet['H9'].value = periode
        # H11 : Total Jam Kerja 
        sheet['H11'].value = tmp_summary['monthly_billable_hours'].iloc[0]  
        
        # H9 : Periode Gaji 
        sheet['I14'].value = tmp_summary['basic_salary'].iloc[0]
        # H11 : Total Jam Kerja 
        sheet['I15'].value = tmp_summary['uang_makan'].iloc[0]          

        # G14 : Makan & Transport 
        sheet['G14'].value = tmp_summary['monthly_billable_hours'].iloc[0]   
        # G15 : Makan & Transport 
        sheet['G15'].value = tmp_summary['total_meal_days'].iloc[0]  
        # K23 : BPJSTK 
        sheet['K23'].value = tmp_summary['bpjstk_deduction'].iloc[0]  
        # K24 : Pajak
        sheet['K24'].value = tmp_summary['tax_deduction'].iloc[0]  
        
        
    sheet_to_delete = wb['tmp']
    wb.remove(sheet_to_delete)    
    
    # PAYROLL SHEET
    wb.create_sheet('Payroll')
    payroll_sheet = wb['Payroll']
    
    payroll_df = summary_salary_df[['norek','nama','net_salary']]
    payroll_df.columns = ['NOMOR REK','NAMA','JUMLAH']
    payroll_df.reset_index()
    payroll_df = payroll_df.rename(columns={'index': 'NOMOR'})
    payroll_df['NOMOR'] = payroll_df.index + 1
    columns_order = ['NOMOR'] + [col for col in payroll_df.columns if col != 'NOMOR']
    payroll_df = payroll_df[columns_order]
    
    for r_idx, row in enumerate(dataframe_to_rows(payroll_df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                payroll_sheet.cell(row=r_idx, column=c_idx, value=value)
    
    rupiah_style = NamedStyle(name='rupiah_style', number_format='Rp #,##0')
    for row in payroll_sheet.iter_rows(min_col=4, max_col=4, min_row=2, max_row=999):
        for cell in row:
            cell.style = rupiah_style
            
    for col in payroll_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        
        for cell in col:
            try:
                # Find the length of the cell value, considering it's a string
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length
        payroll_sheet.column_dimensions[column_letter].width = adjusted_width
    # Define center alignment
    center_aligned_text = Alignment(horizontal='center', vertical='center')

    # Apply center alignment to all cells in the first row
    for cell in payroll_sheet[1]:  # sheet[1] refers to the first row
        cell.alignment = center_aligned_text
    

    wb.remove(payroll_sheet)
    wb._sheets.insert(0, payroll_sheet)
    
    # Detail Jam Kerja
    wb.create_sheet('Rincian Jam Kerja')
    working_hours_sheet = wb['Rincian Jam Kerja']
    
    
    for r_idx, row in enumerate(dataframe_to_rows(working_hours_df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            working_hours_sheet.cell(row=r_idx, column=c_idx, value=value)
    
    wb.remove(working_hours_sheet)
    wb._sheets.insert(1, working_hours_sheet)
    

    # # Detail Perhitungan Gaji
    # wb.create_sheet('Detail Perhitungan Gaji')
    # salary_detail_sheet = wb['Detail Perhitungan Gaji']
    
    
    # for r_idx, row in enumerate(dataframe_to_rows(detail_salary_df, index=False, header=True), start=1):
    #     for c_idx, value in enumerate(row, start=1):
    #         salary_detail_sheet.cell(row=r_idx, column=c_idx, value=value)
    
    # wb.remove(salary_detail_sheet)
    # wb._sheets.insert(2, salary_detail_sheet)
    
    
    #Gaji Harian Template
    gaji_harian_sheet = wb['tmp2']
    
    gaji_harian_df = report_gaji_harian(detail_salary_df)
    for r_idx, row in enumerate(dataframe_to_rows(gaji_harian_df, index=False, header=False), start=8):
        for c_idx, value in enumerate(row, start=2):
            gaji_harian_sheet.cell(row=r_idx, column=c_idx, value=value)
    gaji_harian_sheet.title = 'Gaji Harian'
    wb.remove(gaji_harian_sheet)
    wb._sheets.insert(1, gaji_harian_sheet)
    
    
    wb.save(file_output)
    
    return file_output
    