import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import streamlit as st
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def report_by_project(df):
    # Ensure the 'tanggal' column is in datetime format
    df['tanggal'] = pd.to_datetime(df['tanggal'], format='%Y-%m-%d')

    # Group by 'Project', 'nama', and 'jabatan' and pivot the table
    pivot_df = df.pivot_table(index=['nik','project', 'nama', 'jabatan','uang_makan','basic_salary'], 
                              columns=df['tanggal'].dt.day, 
                              values='billable_hours', 
                              aggfunc='sum').reset_index()
    
    # Create a list of all days (1 to 31)
    all_days = [str(i) for i in range(1, 32)]
    
    # Reindex the pivot table to ensure all days are present
    pivot_df = pivot_df.reindex(columns=['nik','project', 'nama', 'jabatan','uang_makan','basic_salary'] + list(range(1, 32)), fill_value=0)
    
    # Rename columns
    pivot_df.columns = ['nik','project', 'nama', 'jabatan','uang_makan','basic_salary'] + all_days

    # Calculate the total working hours (TOTAL JAM KERJA)
    pivot_df['TOTAL JAM KERJA'] = pivot_df.loc[:, '1':'31'].sum(axis=1)

    # Calculate the total working days (TOTAL HARI KERJA)
    pivot_df['TOTAL HARI KERJA'] = (pivot_df.loc[:, '1':'31'] > 5).sum(axis=1)

    # Calculate meal allowance (UANG MAKAN)
    pivot_df['BASIC/JAM'] = pivot_df['basic_salary']
    pivot_df['UANG MAKAN'] = pivot_df['uang_makan']

    pivot_df['TOTAL UANG MAKAN'] = pivot_df['UANG MAKAN'] * pivot_df['TOTAL HARI KERJA']

    # Calculate basic salary per hour (BASIC/JAM)


    # Calculate total basic salary (TOTAL BASIC)
    pivot_df['TOTAL BASIC'] = pivot_df['TOTAL JAM KERJA'] * pivot_df['BASIC/JAM']

    # Calculate subtotal (SUB TOTAL)
    pivot_df['SUB TOTAL'] = pivot_df['TOTAL BASIC'] + pivot_df['TOTAL UANG MAKAN']

    # Select and order the columns for the final DataFrame
    final_df = pivot_df[['project', 'nama', 'jabatan'] + all_days + 
                        ['TOTAL JAM KERJA', 'TOTAL HARI KERJA', 'UANG MAKAN', 
                         'BASIC/JAM','TOTAL UANG MAKAN', 'TOTAL BASIC', 'SUB TOTAL']]

    
    
    
    return final_df


def generate_report(report_df,periode):
    
    
    filename = periode.replace(" ", "_")
    file_output = Path('output/report_project_'+filename+".xlsx")
    
    project_list = report_df['project'].unique()
    wb = openpyxl.load_workbook(Path('template/template_report.xlsx'))
    
    for project in project_list:
        sheet = wb.copy_worksheet(wb.active)
        sheet.title = project
        
        tmp_detail = report_df[report_df['project']==str(project)]
        tmp_detail = tmp_detail.drop('project', axis=1)
        
        sheet['M1'].value = project

        start_row = 6
        start_col = 2  # Column B
        for r_idx, row in enumerate(dataframe_to_rows(tmp_detail, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                sheet.cell(row=r_idx, column=c_idx, value=value)
    
    
    sheet_to_delete = wb['tmp']
    wb.remove(sheet_to_delete)   
    wb.save(file_output)
    
    return file_output
