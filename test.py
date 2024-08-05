import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import calendar


# Generate dates for July 2024
dates = pd.date_range(start="2024-07-01", end="2024-07-31", freq='D')

# Create sample data
data = {
    'tanggal': [date.strftime('%Y-%m-%d') for date in dates],
    'jam masuk': ['08:00'] * len(dates),
    'jam kerja': ['16:00'] * len(dates),
    'total jam in 1 month': [8] * len(dates)
}

# Create DataFrame
df = pd.DataFrame(data)

# Create a new Workbook
#wb = Workbook()
wb = openpyxl.load_workbook('/Users/stefkim/Projects/Daytonna/daytonna_mps/template_gaji.xlsx')
ws = wb.active

# Write DataFrame to specific cells
# Define the starting row and column
start_row = 5
start_col = 1  # Column A

# Write DataFrame to the worksheet
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
    for c_idx, value in enumerate(row, start=start_col):
        ws.cell(row=r_idx, column=c_idx, value=value)

# H4 : Employee Number
ws['H4'].value = 'Stef'
# H5 : Name 
ws['H5'].value = 'Stef' 
# H6 : Jabatan 
ws['H6'].value = 'Stef' 
# H7 : Status Keluarga 
ws['H7'].value = 'Stef' 
# H8 : NPWP 
ws['H8'].value = 'Stef' 
# H9 : Periode Gaji 
ws['H9'].value = 'Stef'
# H11 : Total Jam Kerja 
ws['H11'].value = 'Stef'  


# Save the workbook
wb.save('output.xlsx')
